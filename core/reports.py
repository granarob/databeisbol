import io
from django.http import HttpResponse
from django.db.models import Sum, F
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from .models import StatsBatting, StatsPitching, Team, League, Season, Category

@api_view(['GET'])
@permission_classes([AllowAny])
def export_batting_leaders_pdf(request):
    """Genera un PDF con los líderes de bateo."""
    season_id = request.query_params.get('season')
    category_id = request.query_params.get('category')
    
    qs = StatsBatting.objects.all()
    title_suffix = ""
    
    if season_id:
        qs = qs.filter(game__season_id=season_id)
        s = Season.objects.filter(id=season_id).first()
        if s: title_suffix += f" - {s.name}"
    if category_id:
        qs = qs.filter(game__category_id=category_id)
        c = Category.objects.filter(id=category_id).first()
        if c: title_suffix += f" ({c.name})"

    agg = (
        qs
        .values('player__first_name', 'player__last_name', 'team__name')
        .annotate(
            ab=Sum('ab'), h=Sum('h'), hr=Sum('hr'), rbi=Sum('rbi'), bb=Sum('bb')
        )
        .filter(ab__gt=0)
    )

    results = []
    for row in agg:
        ab = row['ab'] or 1
        h = row['h'] or 0
        row['avg'] = round(h / ab, 3)
        results.append(row)

    # Ordenar por AVG por defecto para el reporte
    results.sort(key=lambda x: x['avg'], reverse=True)
    results = results[:20]  # Top 20

    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Título
    elements.append(Paragraph(f"Reporte de Líderes de Bateo{title_suffix}", styles['Title']))
    elements.append(Spacer(1, 12))

    # Tabla
    data = [['RK', 'JUGADOR', 'EQUIPO', 'AB', 'H', 'HR', 'RBI', 'AVG']]
    for i, r in enumerate(results):
        name = f"{r['player__first_name']} {r['player__last_name']}"
        data.append([
            i + 1,
            name[:20],
            r['team__name'][:15],
            r['ab'],
            r['h'],
            r['hr'],
            r['rbi'],
            f"{r['avg']:.3f}"
        ])

    t = Table(data, colWidths=[30, 140, 120, 40, 40, 40, 40, 50])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(t)
    
    doc.build(elements)
    
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

@api_view(['GET'])
@permission_classes([AllowAny])
def export_standings_excel(request):
    """Genera un Excel con la tabla de posiciones."""
    season_id = request.query_params.get('season')
    
    wb = Workbook()
    wb.remove(wb.active) # Borrar la hoja por defecto

    # Obtener categorías de la temporada
    categories = Category.objects.all()
    if season_id:
        categories = categories.filter(league__season__id=season_id).distinct()
    
    for cat in categories:
        ws = wb.create_sheet(title=cat.name[:30])
        
        # Header
        headers = ['Equipo', 'JJ', 'JG', 'JP', 'JE', 'AVG', 'Dif']
        ws.append(headers)
        
        # Estilo header
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E63946", end_color="E63946", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Equipos de la categoría
        teams = Team.objects.filter(category=cat)
        if season_id:
            teams = teams.filter(season_id=season_id)
        
        # Ordenar por puntos (won)
        ordered_teams = sorted(teams, key=lambda x: x.won, reverse=True)
        
        if ordered_teams:
            max_wins = ordered_teams[0].won
            for t in ordered_teams:
                jj = t.won + t.lost + t.tied
                avg = round(t.won / (t.won + t.lost), 3) if (t.won + t.lost) > 0 else 0
                dif = max_wins - t.won
                ws.append([t.name, jj, t.won, t.lost, t.tied, avg, dif])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = "Tabla_de_Posiciones.xlsx"
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response
